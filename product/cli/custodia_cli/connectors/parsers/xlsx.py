"""
Parser XLSX → testo tabellare deterministico via openpyxl.

Output: per ogni sheet una sezione introdotta da ``## Sheet: <nome>``, seguita
dalle righe in formato tab-separated. Gli LLM downstream digeriscono molto bene
TSV (più di JSON nidificato per dati flat).

Apertura in ``read_only=True, data_only=True``: legge i valori cached delle
formule, non le formule stesse; safe per file grandi (streaming row-by-row).

Anche i Google Sheets nativi passano qui: il connettore Drive li esporta
come XLSX.

Security: XLSX è un container ZIP con XML interno. Mettiamo in atto due
difese: (1) ``defusedxml.defuse_stdlib()`` applicato all'import del package
``parsers`` (protegge parser XML stdlib da XXE/billion-laughs); (2) controllo
dimensione *uncompressed* dell'archivio prima di aprirlo (``check_zip_
uncompressed_size``) per bloccare zip-bomb classiche.
"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from custodia_cli.connectors.base import ParserError


def _stringify_cell(value: object) -> str:
    """Converte un valore di cella in stringa stabile; ``None`` → stringa vuota."""
    if value is None:
        return ""
    return str(value)


def parse_xlsx(content: bytes | Path) -> str:
    """Estrae tutti gli sheet di un workbook XLSX come testo tab-separated.

    Args:
        content: bytes del file o ``Path`` a un XLSX su disco.

    Returns:
        Testo concatenato. Ogni sheet prefissato da ``## Sheet: <nome>`` e
        righe separate da newline, celle da tab.

    Raises:
        ParserError: se il file non è un XLSX valido.
    """
    # Difesa zip-bomb prima di passare il file a openpyxl. Import locale per
    # evitare ciclo con il package ``parsers``.
    from custodia_cli.connectors.parsers import check_zip_uncompressed_size

    check_zip_uncompressed_size(content)

    stream: io.BytesIO | Path
    if isinstance(content, Path):
        stream = content
    else:
        stream = io.BytesIO(content)

    try:
        workbook = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise ParserError(f"XLSX non valido: {exc}") from exc
    except Exception as exc:  # noqa: BLE001
        raise ParserError(f"Errore inatteso aprendo XLSX: {exc}") from exc

    parts: list[str] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        parts.append(f"## Sheet: {sheet_name}")
        for row in worksheet.iter_rows(values_only=True):
            line = "\t".join(_stringify_cell(c) for c in row)
            # Saltiamo righe completamente vuote per ridurre rumore al LLM.
            if line.strip("\t").strip():
                parts.append(line)
        parts.append("")  # blank line fra sheet

    workbook.close()
    return "\n".join(parts).rstrip()


__all__ = ["parse_xlsx"]
