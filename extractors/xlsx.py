"""Extractor XLSX. openpyxl + euristica 'trova tabella vera' per fogli PMI."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from ._base import ExtractionResult, Extractor

logger = logging.getLogger(__name__)

# Numero minimo di celle contigue non vuote in una riga per considerarla "header".
_MIN_HEADER_CELLS = 3
# Tronchiamo i fogli molto grandi nell'output markdown.
_MAX_ROWS_PER_TABLE = 1000


class XlsxExtractor(Extractor):
    """XLSX → markdown.

    Per ogni foglio non vuoto:
      - H2 con nome foglio
      - euristica: prima riga con ``>=_MIN_HEADER_CELLS`` celle contigue → header
      - lettura fino a prima riga interamente vuota → tabella markdown
      - tutto ciò che non rientra nella tabella → sezione "Annotazioni libere" come bullet

    Macro VBA / pivot dinamiche / link esterni: quality = 0.3 + warning.
    """

    name = "xlsx"
    mimes = ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]
    extensions = [".xlsx", ".xlsm"]

    def extract(self, file_path: Path) -> ExtractionResult:
        warnings: list[str] = []
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError:
            warnings.append("openpyxl non installato: estrazione vuota")
            return ExtractionResult(
                markdown="", metadata={"engine": "none"}, warnings=warnings, quality=0.0
            )

        try:
            # ``data_only=True``: ricava il valore calcolato delle formule (se presente nel file).
            wb = load_workbook(filename=str(file_path), data_only=True, read_only=False)
        except Exception as exc:
            warnings.append(f"openpyxl non riesce ad aprire il file: {exc}")
            return ExtractionResult(
                markdown="", metadata={"engine": "openpyxl"}, warnings=warnings, quality=0.0
            )

        has_macros = bool(getattr(wb, "vba_archive", None))
        has_external_links = bool(getattr(wb, "_external_links", None))
        sections: list[str] = []
        sheets_meta: list[dict[str, Any]] = []
        for sheet in wb.worksheets:
            grid = _read_sheet(sheet)
            if not _grid_has_content(grid):
                continue
            has_pivots = bool(getattr(sheet, "_pivots", None))
            section, info = _render_sheet(sheet.title, grid, has_pivots=has_pivots)
            sections.append(section)
            sheets_meta.append({"name": sheet.title, **info})

        markdown = "\n\n".join(sections).strip() or "_(workbook vuoto)_"
        metadata: dict[str, Any] = {
            "engine": "openpyxl",
            "sheets": sheets_meta,
            "sheets_count": len(sheets_meta),
        }
        quality = 0.85 if sheets_meta else 0.1

        if has_macros:
            warnings.append("Workbook contiene macro VBA (.xlsm): output potrebbe essere parziale")
            metadata["has_macros"] = True
            quality = min(quality, 0.3)
        if has_external_links:
            warnings.append("Workbook ha link a fonti esterne: valori potrebbero essere stale")
            metadata["has_external_links"] = True
            quality = min(quality, 0.3)
        if any(s.get("has_pivots") for s in sheets_meta):
            warnings.append(
                "Almeno un foglio contiene pivot dinamiche: solo cache statica esportata"
            )
            quality = min(quality, 0.3)
        if any(s.get("truncated") for s in sheets_meta):
            warnings.append(f"Tabella molto grande: troncata a {_MAX_ROWS_PER_TABLE} righe")

        return ExtractionResult(
            markdown=markdown,
            metadata=metadata,
            warnings=warnings,
            quality=quality,
        )


def _read_sheet(sheet) -> list[list]:
    """Legge l'intero foglio come matrice di valori (None per celle vuote)."""
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _grid_has_content(grid: list[list]) -> bool:
    return any(any(c not in (None, "") for c in row) for row in grid)


def _render_sheet(name: str, grid: list[list], has_pivots: bool) -> tuple[str, dict]:
    """Renderizza un foglio in markdown + dict di metadati per la diagnostica."""
    parts: list[str] = [f"## {name}"]
    header_row_idx, header_cols = _find_header(grid)
    info: dict[str, Any] = {"has_pivots": has_pivots, "rows": len(grid)}

    if header_row_idx is None:
        # Nessun header riconoscibile: rendi tutto come annotazioni libere.
        parts.append("### Annotazioni libere\n\n" + _free_cells(grid))
        info["mode"] = "free-only"
        return "\n\n".join(parts), info

    # Trova la fine della tabella (prima riga interamente vuota dopo l'header).
    end_idx = len(grid)
    for i in range(header_row_idx + 1, len(grid)):
        if all(c in (None, "") for c in grid[i]):
            end_idx = i
            break

    data_rows = grid[header_row_idx + 1 : end_idx]
    truncated = False
    if len(data_rows) > _MAX_ROWS_PER_TABLE:
        data_rows = data_rows[:_MAX_ROWS_PER_TABLE]
        truncated = True
    info["truncated"] = truncated

    # Restringi alle colonne dell'header.
    col_start, col_end = header_cols
    header_cells = [_cell(c) for c in grid[header_row_idx][col_start:col_end]]
    body_cells = [[_cell(c) for c in (row[col_start:col_end] if row else [])] for row in data_rows]
    parts.append(_table_md(header_cells, body_cells))

    # Annotazioni libere: tutto ciò che è prima dell'header o dopo end_idx, o fuori dalle colonne.
    free = _collect_free_cells(grid, header_row_idx, end_idx, col_start, col_end)
    if free:
        parts.append("### Annotazioni libere\n\n" + free)
    info["mode"] = "table+free" if free else "table"
    info["table_rows"] = len(body_cells)
    info["table_cols"] = col_end - col_start
    return "\n\n".join(parts), info


def _find_header(grid: list[list]) -> tuple[int | None, tuple[int, int]]:
    """Trova la prima riga con >= _MIN_HEADER_CELLS celle contigue non vuote.

    Ritorna ``(row_index, (col_start, col_end))`` oppure ``(None, (0, 0))``.
    """
    for i, row in enumerate(grid):
        if not row:
            continue
        # Cerca il run contiguo più lungo di celle non vuote.
        best_start, best_len = -1, 0
        cur_start, cur_len = -1, 0
        for j, val in enumerate(row):
            if val not in (None, ""):
                if cur_len == 0:
                    cur_start = j
                cur_len += 1
                if cur_len > best_len:
                    best_start, best_len = cur_start, cur_len
            else:
                cur_len = 0
        if best_len >= _MIN_HEADER_CELLS:
            return i, (best_start, best_start + best_len)
    return None, (0, 0)


def _cell(value) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", " ").strip()


def _table_md(header: list[str], body: list[list[str]]) -> str:
    if not header:
        return ""
    sep = ["---"] * len(header)
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(sep) + " |"]
    for row in body:
        # Normalizza lunghezza.
        if len(row) < len(header):
            row = row + [""] * (len(header) - len(row))
        elif len(row) > len(header):
            row = row[: len(header)]
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def _collect_free_cells(
    grid: list[list], header_row: int, end_row: int, col_start: int, col_end: int
) -> str:
    """Bullet list di celle non vuote che cadono fuori dalla tabella principale."""
    bullets: list[str] = []
    for i, row in enumerate(grid):
        if not row:
            continue
        for j, val in enumerate(row):
            if val in (None, ""):
                continue
            in_table = (header_row <= i < end_row) and (col_start <= j < col_end)
            if not in_table:
                bullets.append(f"- `{_excel_addr(i, j)}`: {_cell(val)}")
    return "\n".join(bullets)


def _excel_addr(row: int, col: int) -> str:
    """(0-based) → notazione Excel (A1, AA12, ...)."""
    letters = ""
    n = col
    while True:
        letters = chr(ord("A") + n % 26) + letters
        n = n // 26 - 1
        if n < 0:
            break
    return f"{letters}{row + 1}"


def _free_cells(grid: list[list]) -> str:
    return (
        _collect_free_cells(grid, header_row=-1, end_row=-1, col_start=-1, col_end=-1)
        or "_(nessuna cella significativa)_"
    )
